"""Friendly CSV, JSON and Excel import helpers for FlowGuard.

The product accepts tables created by non-technical users. Column names are
mapped to the stable internal schema without changing the source file.
"""

from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO, StringIO
import csv
import re
from typing import Any, Iterable

from openpyxl import load_workbook


INTERNAL_FIELDS = (
    "id",
    "client_name",
    "phone",
    "email",
    "service",
    "scheduled_at",
    "status",
    "amount",
    "notes",
    "source",
    "owner",
)

COLUMN_ALIASES = {
    "id": {"id", "ид", "номер", "номер заявки", "id заявки", "request id"},
    "client_name": {
        "client name",
        "client_name",
        "customer",
        "customer name",
        "name",
        "клиент",
        "имя",
        "имя клиента",
        "фио",
    },
    "phone": {
        "phone",
        "phone number",
        "telephone",
        "mobile",
        "телефон",
        "номер телефона",
        "мобильный",
    },
    "email": {"email", "e mail", "mail", "почта", "электронная почта"},
    "service": {
        "service",
        "product",
        "order",
        "услуга",
        "товар",
        "заказ",
        "предмет заявки",
    },
    "scheduled_at": {
        "scheduled at",
        "scheduled_at",
        "date",
        "datetime",
        "appointment",
        "дата",
        "дата записи",
        "время записи",
        "дата и время",
    },
    "status": {"status", "state", "статус", "состояние"},
    "amount": {"amount", "sum", "price", "cost", "сумма", "стоимость", "цена"},
    "notes": {
        "notes",
        "note",
        "comment",
        "comments",
        "message",
        "комментарий",
        "примечание",
        "сообщение",
    },
    "source": {"source", "channel", "источник", "канал"},
    "owner": {
        "owner",
        "manager",
        "responsible",
        "employee",
        "ответственный",
        "менеджер",
        "сотрудник",
    },
}


def _normalise_header(value: Any) -> str:
    text = str(value or "").strip().lower().replace("ё", "е")
    text = re.sub(r"[_\-]+", " ", text)
    return " ".join(text.split())


ALIAS_TO_FIELD = {
    _normalise_header(alias): field
    for field, aliases in COLUMN_ALIASES.items()
    for alias in aliases | {field}
}


@dataclass(frozen=True)
class ImportResult:
    records: list[dict[str, Any]]
    column_mapping: dict[str, str]
    ignored_columns: list[str]
    source_format: str
    sheet_name: str | None = None

    def as_info(self) -> dict[str, Any]:
        return {
            "source_format": self.source_format,
            "sheet_name": self.sheet_name,
            "column_mapping": self.column_mapping,
            "ignored_columns": self.ignored_columns,
        }


def map_rows(
    rows: Iterable[dict[str, Any]],
    headers: Iterable[Any],
    *,
    source_format: str,
    sheet_name: str | None = None,
) -> ImportResult:
    header_names = [str(header or "").strip() for header in headers]
    mapping = {
        header: ALIAS_TO_FIELD[_normalise_header(header)]
        for header in header_names
        if _normalise_header(header) in ALIAS_TO_FIELD
    }
    if not mapping:
        raise ValueError(
            "Не удалось распознать столбцы. Нужен хотя бы один столбец вроде "
            "«Клиент», «Телефон», «Почта», «Сумма», «Статус» или «Комментарий»."
        )

    records: list[dict[str, Any]] = []
    for row in rows:
        if not any(value not in (None, "") for value in row.values()):
            continue
        result: dict[str, Any] = {}
        for original, canonical in mapping.items():
            value = row.get(original)
            if canonical not in result or result[canonical] in (None, ""):
                result[canonical] = value
        records.append(result)

    if not records:
        raise ValueError("В таблице нет строк с данными.")

    ignored = [header for header in header_names if header and header not in mapping]
    return ImportResult(records, mapping, ignored, source_format, sheet_name)


def records_from_csv(text: str) -> ImportResult:
    reader = csv.DictReader(StringIO(text))
    if not reader.fieldnames:
        raise ValueError("В CSV-файле отсутствует строка с названиями столбцов.")
    return map_rows(reader, reader.fieldnames, source_format="csv")


def records_from_json(value: Any) -> ImportResult:
    records = value.get("records") if isinstance(value, dict) else value
    if not isinstance(records, list) or not all(isinstance(row, dict) for row in records):
        raise ValueError("JSON должен содержать массив объектов или поле records с таким массивом.")
    headers: list[str] = []
    for row in records:
        for key in row:
            if key not in headers:
                headers.append(key)
    return map_rows(records, headers, source_format="json")


def records_from_xlsx(payload: bytes) -> ImportResult:
    try:
        workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Не удалось открыть Excel-файл. Проверьте, что это корректный .xlsx.") from exc

    sheet = workbook.active
    values = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(values)
    except StopIteration as exc:
        raise ValueError("Excel-файл пуст.") from exc

    headers = [str(value or "").strip() for value in raw_headers]
    rows = (dict(zip(headers, row, strict=False)) for row in values)
    return map_rows(rows, headers, source_format="xlsx", sheet_name=sheet.title)
