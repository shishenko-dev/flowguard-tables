"""Локальный сервер FlowGuard Таблицы.

Run with Python 3.11+: python app.py
Open http://127.0.0.1:8787
"""

from __future__ import annotations

import argparse
import csv
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO, StringIO
import json
import os
from pathlib import Path
import sys
from threading import Timer
from urllib.parse import urlparse
import webbrowser

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from flowguard import IntentClassifier, analyze_records
from flowguard.importers import records_from_csv, records_from_json, records_from_xlsx
from flowguard.storage import RunStore


ROOT = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
STATIC = ROOT / "static"
SAMPLE = ROOT / "data" / "sample_records.csv"
custom_data_dir = os.getenv("FLOWGUARD_DATA_DIR")
if custom_data_dir:
    user_data = Path(custom_data_dir).expanduser().resolve()
    user_data.mkdir(parents=True, exist_ok=True)
    store_path = user_data / "flowguard.sqlite3"
elif getattr(sys, "frozen", False):
    user_data = Path(os.getenv("LOCALAPPDATA", Path.home())) / "FlowGuardTables"
    user_data.mkdir(parents=True, exist_ok=True)
    store_path = user_data / "flowguard.sqlite3"
else:
    store_path = ROOT / "data" / "flowguard.sqlite3"
STORE = RunStore(store_path)
CLASSIFIER = IntentClassifier()
MAX_BODY_BYTES = 15_000_000


def read_sample() -> list[dict[str, str]]:
    with SAMPLE.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def records_to_csv(records: list[dict[str, object]]) -> bytes:
    columns = [
        "id",
        "client_name",
        "phone",
        "email",
        "service",
        "scheduled_at",
        "status",
        "amount",
        "priority",
        "risk_score",
        "recommended_action",
        "intent",
        "notes",
        "source",
        "owner",
        "flags",
        "reasons",
    ]
    buffer = StringIO()
    writer = csv.DictWriter(buffer, fieldnames=columns)
    writer.writeheader()
    for record in records:
        row = {column: record.get(column, "") for column in columns}
        intent = record.get("intent", {})
        row["intent"] = intent.get("label", "") if isinstance(intent, dict) else ""
        row["flags"] = "; ".join(record.get("flags", []))
        row["reasons"] = "; ".join(record.get("reasons", []))
        for column, value in row.items():
            if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
                row[column] = "'" + value
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8-sig")


def records_to_xlsx(analysis: dict[str, object]) -> bytes:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Сводка"
    summary = analysis.get("summary", {})
    summary_rows = [
        ("Показатель", "Значение"),
        ("Всего строк", summary.get("total", 0)),
        ("Требуют внимания", summary.get("attention", 0)),
        ("Критические", summary.get("critical", 0)),
        ("Группы повторов", summary.get("duplicate_groups", 0)),
        ("Аномальные суммы", summary.get("anomalies", 0)),
    ]
    for row in summary_rows:
        summary_sheet.append(row)

    columns = [
        "id", "client_name", "phone", "email", "service", "scheduled_at",
        "status", "amount", "priority", "risk_score", "recommended_action",
        "notes", "source", "owner", "flags", "reasons",
    ]
    labels = [
        "ID", "Клиент", "Телефон", "Почта", "Услуга", "Дата",
        "Статус", "Сумма", "Приоритет", "Риск", "Рекомендация",
        "Комментарий", "Источник", "Ответственный", "Метки", "Причины",
    ]
    records = analysis.get("records", [])
    for sheet_name, selected in (
        ("Все строки", records),
        ("Требуют внимания", [row for row in records if row.get("priority") != "normal"]),
    ):
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(labels)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="17324D")
        for record in selected:
            values = []
            for column in columns:
                value = record.get(column, "")
                if isinstance(value, list):
                    value = "; ".join(str(item) for item in value)
                if isinstance(value, str) and value.startswith(("=", "+", "-", "@")):
                    value = "'" + value
                values.append(value)
            sheet.append(values)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_cells in sheet.columns:
            width = min(45, max(10, max(len(str(cell.value or "")) for cell in column_cells) + 2))
            sheet.column_dimensions[column_cells[0].column_letter].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


class Handler(SimpleHTTPRequestHandler):
    server_version = "FlowGuardTables/0.1"

    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(STATIC), **kwargs)

    def _json(self, payload: object, status: HTTPStatus = HTTPStatus.OK) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def _error(self, message: str, status: HTTPStatus = HTTPStatus.BAD_REQUEST) -> None:
        self._json({"error": message}, status)

    def do_GET(self) -> None:  # noqa: N802
        path = urlparse(self.path).path
        if path == "/api/health":
            self._json({"status": "ok", "service": "FlowGuard Таблицы", "version": "0.1.0"})
            return
        if path == "/api/sample":
            self._json({"records": read_sample()})
            return
        if path == "/api/model":
            self._json(CLASSIFIER.evaluate())
            return
        if path == "/api/runs":
            self._json({"runs": STORE.list_runs()})
            return
        if path == "/api/report/latest.csv":
            latest = STORE.latest()
            if latest is None:
                self._error("Нет завершённого анализа для выгрузки.", HTTPStatus.NOT_FOUND)
                return
            body = records_to_csv(latest["records"])
            self.send_response(HTTPStatus.OK)
            self.send_header("Content-Type", "text/csv; charset=utf-8")
            self.send_header("Content-Disposition", 'attachment; filename="flowguard-report.csv"')
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        if path == "/api/report/latest.xlsx":
            latest = STORE.latest()
            if latest is None:
                self._error("Нет завершённого анализа для выгрузки.", HTTPStatus.NOT_FOUND)
                return
            body = records_to_xlsx(latest)
            self.send_response(HTTPStatus.OK)
            self.send_header(
                "Content-Type",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            self.send_header("Content-Disposition", 'attachment; filename="flowguard-report.xlsx"')
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return
        if path == "/":
            self.path = "/index.html"
        super().do_GET()

    def do_POST(self) -> None:  # noqa: N802
        if urlparse(self.path).path != "/api/analyze":
            self._error("Адрес запроса не найден.", HTTPStatus.NOT_FOUND)
            return
        try:
            length = int(self.headers.get("Content-Length", "0"))
        except ValueError:
            self._error("Не удалось определить размер файла.")
            return
        if length <= 0 or length > MAX_BODY_BYTES:
            self._error("Размер файла должен быть от 1 байта до 15 МБ.")
            return
        body = self.rfile.read(length)
        content_type = self.headers.get("Content-Type", "").split(";")[0].strip().lower()
        try:
            if content_type in {"text/csv", "application/csv"}:
                imported = records_from_csv(body.decode("utf-8-sig"))
            elif content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                imported = records_from_xlsx(body)
            else:
                payload = json.loads(body.decode("utf-8"))
                imported = records_from_json(payload)
            analysis = analyze_records(imported.records, CLASSIFIER)
            analysis["import_info"] = imported.as_info()
            analysis["run_id"] = STORE.save(analysis)
            self._json(analysis, HTTPStatus.CREATED)
        except (UnicodeDecodeError, json.JSONDecodeError, ValueError) as exc:
            self._error(str(exc))
        except Exception as exc:  # Defensive boundary for a demo service.
            print(f"Unhandled error: {exc}", file=sys.stderr)
            self._error("Проверка безопасно остановлена. Подробности сохранены в журнале программы.", HTTPStatus.INTERNAL_SERVER_ERROR)


def main() -> None:
    parser = argparse.ArgumentParser(description="Запустить FlowGuard Таблицы")
    parser.add_argument("--host", default="127.0.0.1", help="Bind address (default: local computer only)")
    parser.add_argument("--port", type=int, default=8787, help="TCP port (default: 8787)")
    parser.add_argument("--open-browser", action="store_true", help="Открыть программу в браузере")
    args = parser.parse_args()
    server = ThreadingHTTPServer((args.host, args.port), Handler)
    address = f"http://{args.host}:{args.port}"
    print(f"FlowGuard Таблицы запущен: {address}")
    print("Чтобы остановить программу, закройте это окно или нажмите Ctrl+C.")
    if args.open_browser:
        Timer(0.6, lambda: webbrowser.open(address)).start()
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nFlowGuard Таблицы остановлен.")
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
